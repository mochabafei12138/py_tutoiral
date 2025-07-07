import openpyxl

# 1.获取工作簿对象
workbook = openpyxl.load_workbook(r'data/阿里巴巴2020年股票数据.xlsx')
print(workbook)

# 2.获取工作表对象
# a.获取所有工作表名称
print(workbook.sheetnames)
# b.获取当前处于活跃状态的工作表对象
# 注意：读取excel文件的时候，可以不关闭文件，但是获取的结果可能不准确
sheet1 = workbook.active  # 最近一次关闭工作簿之后，选中的工作表
print(sheet1)

# c.通过工作表名称获取工作表对象，语法：wb['工作表名称']
sheet2 = workbook['Sheet1']
print(sheet2)

# d.通过索引获取工作表对象，语法：wb.worksheets[index]
# 获取所有的工作表对象，返回一个列表
print(workbook.worksheets)
sheet3 = workbook.worksheets[2]
print(sheet3)

# 3.获取工作表的行数和列数【范围】
print(sheet1.dimensions)   # A1:G255
print(sheet1.max_row,sheet1.max_column)   # 255 7

# 4.获取单元格对象
# a.sheet['B5']
cell1 = sheet1['A1']
print(cell1)

# b.sheet.cell(row,col)
'''
openpyxl
row:
    Excel: 1 2 3 4 5 6....
    Python:1 2 3 4 5 6....
col:
    Excel: A B C D....
    Python:1 2 3 4.....   
'''
cell2 = sheet1.cell(1,2)
print(cell2)

# 练习：C5
print(sheet1['C5'])
print(sheet1.cell(5,3))

# 5.获取单元格的值
print(cell1.value)
print(cell2.value)

# 6.获取指定行或指定列，返回一个元组，其中的元素是cell对象
# a.获取行
# sheet[row]:row和Excel中是对应的
row = sheet1[1]
print(row)
# for cell in row:
#     print(cell.value)
row_data = [cell.value for cell in row]
print(row_data)

# b.获取列
# sheet[col]:col和Excel中是对应的
col = sheet1['C']
print(col)
col_data = [cell.value for cell in col]
print(col_data)

# c.获取多行或多列，返回一个二维元组，其中的元素是一行的单元格对象的元组
rows = sheet1['A2:D10']     # 闭区间
print(rows)
for row in rows:
    for cell in row:
        print(cell.value)