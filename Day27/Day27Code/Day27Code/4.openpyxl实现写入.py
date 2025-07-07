import openpyxl

# 1.创建工作簿
workbook = openpyxl.Workbook()
# 注意：区别于xlwt,当创建完工作簿之后，会自动创建一个默认的工作表，默认名为Sheet
print(workbook.sheetnames)   # ['Sheet']

# 给工作表重命名
sheet = workbook['Sheet']
sheet.title = '用户表'
print(workbook.sheetnames)   # ['用户表']

# 2.创建表
# 注意：创建表的时候，如果没有指明名称，则默认为Sheet,Sheet1......
workbook.create_sheet()
print(workbook.sheetnames)
workbook.create_sheet()
print(workbook.sheetnames)
workbook.create_sheet()
print(workbook.sheetnames)
workbook.create_sheet('成绩表')
print(workbook.sheetnames)

# 3.删除表
sheet1 = workbook['Sheet1']
workbook.remove(sheet1)
print(workbook.sheetnames)

# 4.复制表
sheet2 = workbook['用户表']
workbook.copy_worksheet(sheet2)
print(workbook.sheetnames)
# 给复制之后的表重命名
sheet3 = workbook['用户表 Copy']
sheet3.title = '会员表'
print(workbook.sheetnames)

# 5.修改单元格数据
user_sheet = workbook['用户表']
# 方式一
# sheet['A1'] = 值
user_sheet['A1'] = '姓名'
user_sheet['B1'] = '年龄'
user_sheet['C1'] = '用户编号'

# 方式二
# sheet.cell(row,col,value)
user_sheet.cell(2,1,'小明')
user_sheet.cell(2,2,10)
user_sheet.cell(2,3,'1001')

# 6.追加数据
# 注意：sheet.append()可以向excel中追加一行内容，所以append的参数一般是是可迭代对象【常用列表或元组】
# user_sheet.append(['小花',11,'1002'])

# 注意：后期只要将数据整理成下面二维列表的格式，就可以一次性追加多行
info_list = [
['Tom',13,'1008'],
['Bob',10,'1005'],
['小花',11,'1003'],
['Jack',12,'1007'],
['小花2',15,'1009'],
['小花1',17,'1006'],
['小花3',12,'1004']
]
for info in info_list:
    user_sheet.append(info)

# 7.保存工作簿
workbook.save(r'data/演示写入操作.xlsx')