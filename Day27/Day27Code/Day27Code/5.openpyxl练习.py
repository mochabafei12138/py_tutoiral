import  openpyxl

# 练习一：获取student-score.xlsx中的'成绩表'中涉及哪几个学校，总共有多少个学校
wb = openpyxl.load_workbook(r'data/student-score.xlsx')
# print(wb.sheetnames)
sheet = wb['成绩表']
# 获取c列
c_col = sheet['C'][2:]
# print(c_col)
school_names = set([cell.value for cell in c_col])
print(f'总共涉及到{len(school_names)}个学校，分别是{school_names}')

# 练习二：处理 练习数据.xlsx 文件中的数据，将日期和成交量拆分，保存到一个新的工作簿中，日期和成交量显示两列
wb1 = openpyxl.load_workbook(r'data/练习数据.xlsx')
sheet1 = wb1.active
a_col = sheet1['A']
print(a_col)
# 整理数据
data_list = []
for cell in a_col:
    value = cell.value
    date_num_list = value.split(';')
    print(date_num_list)
    for date_num in date_num_list:
        sublist = date_num.strip().split(":")
        if len(sublist) == 2:
            sublist[1] = sublist[1].strip()
        data_list.append(sublist)
print(data_list)

# 写入
wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2.title = '处理结果'
sheet2.append(['日期','成交量'])
for data in data_list:
    sheet2.append(data)

wb2.save(r'data/练习数据-处理结果.xlsx')

