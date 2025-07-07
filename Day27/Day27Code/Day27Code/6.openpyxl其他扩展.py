from openpyxl import  load_workbook
from openpyxl.styles import Font   # 设置字体样式
from openpyxl.styles import  Side,Border  # 设置边框样式
from openpyxl.styles import  PatternFill,GradientFill  # 设置背景颜色
from openpyxl.styles import  Alignment  # 设置对齐方式

# 1.将公式写入到单元格保存
wb = load_workbook(r'data/student-score1.xlsx')
sheet = wb['成绩表']

# 计算成绩和
sheet['I2'] = '总分'
# 注意：和excel中公式的写法一模一样，只需要给定字符串即可
# sheet['I3'] = '=sum(D3:H3)'
# print(sheet.max_row)

# 求所有学生的总分
for row in range(3,sheet.max_row):
    sheet[f'I{row}'] = f'=sum(D{row}:H{row})'
# 保存
wb.save(r'data/student-score1.xlsx')

# 2.设置单个单元格的字体样式
wb = load_workbook(r'data/student-score1.xlsx')
sheet = wb['成绩表']

cell = sheet['C5']
print(cell.font)
# name='Arial', charset=134, family=None, b=False, i=False, strike=None, outline=None, shadow=None, condense=None, color=None, extend=None, sz=10.0, u=None, vertAlign=None, scheme=None

# 创建字体对象
# font = Font(name='黑体',size=20,bold=True,italic=True,color='66CCFF')  # 颜色只能使用rgb的十六进制表示
# 对一个单元格设置字体样式
# cell.font = font

# 对多个单元格设置字体样式
font = Font(name='黑体',size=15,bold=True,italic=True,color='FF66FF')
cells = sheet[2]
for cell in cells:
    cell.font = font

# 保存
wb.save(r'data/student-score1.xlsx')

# 3.设置边框样式
wb = load_workbook(r'data/student-score1.xlsx')
sheet = wb['成绩表']

# 创建线条对象
side = Side(border_style='double',color='FFCC00')
# 创建边框对象
border = Border(left=side,right=side,top=side,bottom=side)
cells = sheet[2]
for cell in cells:
    cell.border = border
# 保存
wb.save(r'data/student-score1.xlsx')

# 4.设置其他样式
wb = load_workbook(r'data/student-score1.xlsx')
sheet = wb['成绩表']
# a.设置单元格背景色
pattern_fill = PatternFill(fill_type='solid',fgColor='FF6666')
gradient_fill = GradientFill(stop=('FFFFFF','FF3333','000000'))    # 渐变色

# 设置第3行的背景色
cells = sheet[3]
for cell in cells:
    cell.fill = pattern_fill

# 设置第4行的背景色
cells = sheet[5]
for cell in cells:
    cell.fill = gradient_fill

# b.设置文本居中
# horizontal表示水平方向，常用的left center right
# vertical表示垂直方向，常用的top  center  bottom
alignment = Alignment(horizontal='center',vertical='center')
cells = sheet[6]
for cell in cells:
    cell.alignment = alignment

# c.设置行高和列高
print(sheet.row_dimensions)  # 字典
# for r in sheet.row_dimensions:
#     print(r)

# 将第1行的行高设置为30
sheet.row_dimensions[8].height = 30

# 设置第C列的列宽为30
sheet.column_dimensions['C'].width = 30

# 保存
wb.save(r'data/student-score1.xlsx')